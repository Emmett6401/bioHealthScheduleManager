# -*- coding: utf-8 -*-
"""
학생 관리 다이얼로그
"""

from PyQt5.QtWidgets import (QDialog, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QTableWidget, QTableWidgetItem, QLineEdit, QLabel,
                             QComboBox, QMessageBox, QHeaderView, QGroupBox,
                             QGridLayout, QTextEdit, QFileDialog, QProgressDialog)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap, QImage
import sys
import os
import traceback
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from PIL import Image
import io

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database.db_manager import DatabaseManager
from config_db import CODE_PREFIX

# 프로젝트 루트 디렉토리
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PHOTO_DIR = os.path.join(PROJECT_ROOT, 'student_photos')
PHOTO_ORIGINALS = os.path.join(PHOTO_DIR, 'originals')
PHOTO_THUMBNAILS = os.path.join(PHOTO_DIR, 'thumbnails')
DEFAULT_AVATAR = os.path.join(PHOTO_DIR, 'default_avatar.png')


class StudentDialog(QWidget):
    """학생 관리 위젯 (탭용)"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.db = DatabaseManager()
        self.original_code = None  # 수정 시 원본 코드 저장
        self.current_photo_path = None  # 현재 선택된 사진 경로
        self.photo_label = None  # 사진 표시 라벨
        self.init_ui()
        self.load_courses()
        self.load_data()
        
    def init_ui(self):
        """UI 초기화"""
        layout = QVBoxLayout()
        
        # 상단 버튼 그룹
        top_btn_layout = QHBoxLayout()
        
        excel_upload_btn = QPushButton("📂 엑셀 업로드")
        excel_upload_btn.setStyleSheet("background-color: #FF9800; color: white; padding: 10px 20px; font-size: 11pt; font-weight: bold;")
        excel_upload_btn.setMinimumHeight(40)
        excel_upload_btn.clicked.connect(self.upload_excel)
        top_btn_layout.addWidget(excel_upload_btn)
        
        top_btn_layout.addStretch()
        layout.addLayout(top_btn_layout)
        
        # 입력 폼 (사진 + 정보)
        form_group = QGroupBox("학생 정보 등록")
        form_group.setStyleSheet("QGroupBox { font-size: 11pt; font-weight: bold; padding-top: 10px; }")
        main_form_layout = QHBoxLayout()
        
        # 왼쪽: 사진 영역
        photo_widget = QWidget()
        photo_layout = QVBoxLayout()
        photo_layout.setSpacing(5)
        
        self.photo_label = QLabel()
        self.photo_label.setFixedSize(150, 180)
        self.photo_label.setStyleSheet("""
            QLabel {
                border: 2px solid #ccc;
                border-radius: 5px;
                background-color: #f5f5f5;
            }
        """)
        self.photo_label.setAlignment(Qt.AlignCenter)
        self.photo_label.setText("사진 없음")
        photo_layout.addWidget(self.photo_label)
        
        upload_photo_btn = QPushButton("📷 사진 등록")
        upload_photo_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 5px;")
        upload_photo_btn.clicked.connect(self.upload_photo)
        photo_layout.addWidget(upload_photo_btn)
        
        remove_photo_btn = QPushButton("🗑️ 사진 삭제")
        remove_photo_btn.setStyleSheet("padding: 5px;")
        remove_photo_btn.clicked.connect(self.remove_photo)
        photo_layout.addWidget(remove_photo_btn)
        
        photo_layout.addStretch()
        photo_widget.setLayout(photo_layout)
        main_form_layout.addWidget(photo_widget)
        
        # 오른쪽: 정보 입력 폼
        form_layout = QGridLayout()
        form_layout.setSpacing(8)
        form_layout.setVerticalSpacing(8)
        
        # 1행: 코드, 이름, 성별
        form_layout.addWidget(QLabel("코드:"), 0, 0)
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("예: S-001")
        self.code_input.setMaximumWidth(150)
        form_layout.addWidget(self.code_input, 0, 1)
        
        form_layout.addWidget(QLabel("이름:"), 0, 2)
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("홍길동")
        self.name_input.setMaximumWidth(150)
        form_layout.addWidget(self.name_input, 0, 3)
        
        form_layout.addWidget(QLabel("성별:"), 0, 4)
        self.gender_combo = QComboBox()
        self.gender_combo.addItems(["남자", "여자", "선택안함"])
        self.gender_combo.setMaximumWidth(100)
        form_layout.addWidget(self.gender_combo, 0, 5)
        
        # 2행: 생년월일, 휴대폰, 이메일
        form_layout.addWidget(QLabel("생년월일:"), 1, 0)
        self.birth_input = QLineEdit()
        self.birth_input.setPlaceholderText("99.01.12")
        self.birth_input.setMaximumWidth(150)
        form_layout.addWidget(self.birth_input, 1, 1)
        
        form_layout.addWidget(QLabel("휴대폰:"), 1, 2)
        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("010-1234-5678")
        self.phone_input.setMaximumWidth(150)
        form_layout.addWidget(self.phone_input, 1, 3)
        
        form_layout.addWidget(QLabel("이메일:"), 1, 4)
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("example@email.com")
        form_layout.addWidget(self.email_input, 1, 5)
        
        # 3행: 주소
        form_layout.addWidget(QLabel("주소:"), 2, 0)
        self.address_input = QLineEdit()
        self.address_input.setPlaceholderText("주소를 입력하세요")
        form_layout.addWidget(self.address_input, 2, 1, 1, 5)
        
        # 4행: 관심분야, 최종학교
        form_layout.addWidget(QLabel("관심분야:"), 3, 0)
        self.interests_input = QLineEdit()
        self.interests_input.setPlaceholderText("예: 로봇, AI")
        form_layout.addWidget(self.interests_input, 3, 1, 1, 2)
        
        form_layout.addWidget(QLabel("최종학교:"), 3, 3)
        self.education_input = QLineEdit()
        self.education_input.setPlaceholderText("예: 우송대학교/3학년/AI빅데이터학과")
        form_layout.addWidget(self.education_input, 3, 4, 1, 2)
        
        # 5행: 캠퍼스, 배정과정
        form_layout.addWidget(QLabel("지원캠퍼스:"), 4, 0)
        self.campus_input = QLineEdit()
        self.campus_input.setPlaceholderText("예: 우송바이오헬스아카데미")
        form_layout.addWidget(self.campus_input, 4, 1, 1, 2)
        
        form_layout.addWidget(QLabel("배정과정:"), 4, 3)
        self.course_combo = QComboBox()
        self.course_combo.setMaximumWidth(200)
        form_layout.addWidget(self.course_combo, 4, 4, 1, 2)
        
        # 6행: 자기소개
        form_layout.addWidget(QLabel("자기소개:"), 5, 0)
        self.introduction_input = QTextEdit()
        self.introduction_input.setPlaceholderText("자기소개를 입력하세요 (200자 내외)")
        self.introduction_input.setMaximumHeight(60)
        form_layout.addWidget(self.introduction_input, 5, 1, 1, 5)
        
        # 7행: 비고
        form_layout.addWidget(QLabel("비고:"), 6, 0)
        self.notes_input = QTextEdit()
        self.notes_input.setPlaceholderText("비고사항을 입력하세요")
        self.notes_input.setMaximumHeight(40)
        form_layout.addWidget(self.notes_input, 6, 1, 1, 5)
        
        # 폼 레이아웃을 메인 레이아웃에 추가
        form_widget = QWidget()
        form_widget.setLayout(form_layout)
        main_form_layout.addWidget(form_widget, 1)  # stretch factor 1
        
        form_group.setLayout(main_form_layout)
        layout.addWidget(form_group)
        
        # 버튼 그룹
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.add_btn = QPushButton("추가")
        self.add_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px 20px; font-size: 10pt;")
        self.add_btn.setMinimumHeight(35)
        self.add_btn.clicked.connect(self.add_student)
        btn_layout.addWidget(self.add_btn)
        
        self.update_btn = QPushButton("수정")
        self.update_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 8px 20px; font-size: 10pt;")
        self.update_btn.setMinimumHeight(35)
        self.update_btn.clicked.connect(self.update_student)
        btn_layout.addWidget(self.update_btn)
        
        self.delete_btn = QPushButton("삭제")
        self.delete_btn.setStyleSheet("background-color: #f44336; color: white; padding: 8px 20px; font-size: 10pt;")
        self.delete_btn.setMinimumHeight(35)
        self.delete_btn.clicked.connect(self.delete_student)
        btn_layout.addWidget(self.delete_btn)
        
        self.clear_btn = QPushButton("초기화")
        self.clear_btn.setMinimumHeight(35)
        self.clear_btn.clicked.connect(self.clear_form)
        btn_layout.addWidget(self.clear_btn)
        
        layout.addLayout(btn_layout)
        
        # 테이블
        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "코드", "이름", "생년월일", "성별", "휴대폰", "이메일", 
            "관심분야", "최종학교", "캠퍼스", "배정과정", "등록일", "비고"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.cellClicked.connect(self.on_row_selected)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
        
        # 과정 목록 로드
        self.load_courses()
        
    def load_courses(self):
        """과정 목록 로드"""
        try:
            if not self.db.connect():
                return
            
            self.course_combo.clear()
            self.course_combo.addItem("미배정", None)
            
            query = "SELECT code, name FROM courses ORDER BY code"
            courses = self.db.fetch_all(query)
            
            if courses:
                for course in courses:
                    self.course_combo.addItem(f"{course['code']} - {course['name']}", course['code'])
                    
        except Exception as e:
            print(f"과정 목록 로드 오류: {str(e)}")
            
    def upload_excel(self):
        """엑셀 파일 업로드"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "엑셀 파일 선택", "", "Excel Files (*.xlsx *.xls)"
            )
            
            if not file_path:
                return
            
            # 엑셀 파일 읽기
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            
            # 헤더 확인 (첫 번째 행)
            headers = []
            for col in range(1, sheet.max_column + 1):
                headers.append(sheet.cell(1, col).value)
            
            # 데이터 읽기 (2행부터)
            students_data = []
            for row in range(2, sheet.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row, col).value
                    row_data[header] = cell_value
                students_data.append(row_data)
            
            wb.close()
            
            # 데이터 저장
            if not self.db.connect():
                QMessageBox.critical(self, "오류", "데이터베이스 연결 실패")
                return
            
            success_count = 0
            fail_count = 0
            
            for data in students_data:
                try:
                    # 코드 자동 생성
                    code = self.db.get_next_code('students', 'S')
                    
                    # 데이터 매핑
                    name = data.get('이름', '')
                    birth_date = data.get('생년월일(78.01.12)', '')
                    gender = data.get('성별\n(선택)', '')
                    phone = data.get('휴대폰번호', '')
                    email = data.get('이메일', '')
                    address = data.get('주소', '')
                    interests = data.get('관심 있는 분야(2개)', '')
                    education = data.get('최종 학교/학년', '')
                    introduction = data.get('자기소개 (200자 내외)', '')
                    campus = data.get('지원하고자 하는 캠퍼스를 선택하세요', '')
                    
                    if not name or not phone:
                        fail_count += 1
                        continue
                    
                    query = """
                        INSERT INTO students (code, name, birth_date, gender, phone, email, 
                                            address, interests, education, introduction, campus)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    self.db.execute_query(query, (
                        code, name, birth_date, gender, phone, email,
                        address, interests, education, introduction, campus
                    ))
                    success_count += 1
                    
                except Exception as e:
                    print(f"학생 등록 오류 ({name}): {str(e)}")
                    fail_count += 1
            
            QMessageBox.information(
                self, "완료", 
                f"엑셀 업로드 완료\n성공: {success_count}명\n실패: {fail_count}명"
            )
            self.load_data()
            
        except Exception as e:
            error_msg = f"엑셀 업로드 실패: {str(e)}\n\n상세 오류:\n{traceback.format_exc()}"
            print(error_msg)
            QMessageBox.critical(self, "오류", error_msg)
            
    def is_code_duplicate(self, code, exclude_code=None):
        """코드 중복 체크"""
        try:
            if not self.db.connect():
                return False
            
            if exclude_code:
                query = "SELECT COUNT(*) as count FROM students WHERE code = %s AND code != %s"
                result = self.db.fetch_one(query, (code, exclude_code))
            else:
                query = "SELECT COUNT(*) as count FROM students WHERE code = %s"
                result = self.db.fetch_one(query, (code,))
            
            return result and result['count'] > 0
        except Exception as e:
            print(f"코드 중복 체크 오류: {str(e)}")
            return False
            
    def add_student(self):
        """학생 추가"""
        code = self.code_input.text().strip()
        name = self.name_input.text().strip()
        phone = self.phone_input.text().strip()
        
        if not code:
            QMessageBox.warning(self, "경고", "코드를 입력하세요.")
            return
        
        if not name:
            QMessageBox.warning(self, "경고", "이름을 입력하세요.")
            return
        
        if not phone:
            QMessageBox.warning(self, "경고", "휴대폰 번호를 입력하세요.")
            return
        
        # 코드 중복 체크
        if self.is_code_duplicate(code):
            QMessageBox.warning(self, "경고", f"코드 '{code}'는 이미 사용 중입니다.\n다른 코드를 입력하세요.")
            return
        
        birth_date = self.birth_input.text().strip()
        gender = self.gender_combo.currentText()
        email = self.email_input.text().strip()
        address = self.address_input.text().strip()
        interests = self.interests_input.text().strip()
        education = self.education_input.text().strip()
        introduction = self.introduction_input.toPlainText().strip()
        campus = self.campus_input.text().strip()
        course_code = self.course_combo.currentData()
        notes = self.notes_input.toPlainText().strip()
        
        try:
            query = """
                INSERT INTO students (code, name, birth_date, gender, phone, email, 
                                    address, interests, education, introduction, campus, course_code, notes, photo_path)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            self.db.execute_query(query, (
                code, name, birth_date, gender, phone, email,
                address, interests, education, introduction, campus, course_code, notes, self.current_photo_path
            ))
            
            QMessageBox.information(self, "성공", f"학생 {code}가 추가되었습니다.")
            self.clear_form()
            self.load_data()
            
        except Exception as e:
            error_msg = f"추가 실패: {str(e)}\n\n상세 오류:\n{traceback.format_exc()}"
            print(error_msg)
            QMessageBox.critical(self, "오류", error_msg)
            
    def update_student(self):
        """학생 정보 수정"""
        code = self.code_input.text().strip()
        name = self.name_input.text().strip()
        phone = self.phone_input.text().strip()
        
        if not code or not name or not phone:
            QMessageBox.warning(self, "경고", "코드, 이름, 휴대폰 번호를 입력하세요.")
            return
        
        if not self.original_code:
            QMessageBox.warning(self, "경고", "수정할 학생을 먼저 선택하세요.")
            return
        
        # 코드가 변경된 경우 중복 체크
        if code != self.original_code:
            if self.is_code_duplicate(code, self.original_code):
                QMessageBox.warning(self, "경고", f"코드 '{code}'는 이미 사용 중입니다.\n다른 코드를 입력하세요.")
                return
        
        birth_date = self.birth_input.text().strip()
        gender = self.gender_combo.currentText()
        email = self.email_input.text().strip()
        address = self.address_input.text().strip()
        interests = self.interests_input.text().strip()
        education = self.education_input.text().strip()
        introduction = self.introduction_input.toPlainText().strip()
        campus = self.campus_input.text().strip()
        course_code = self.course_combo.currentData()
        notes = self.notes_input.toPlainText().strip()
        
        try:
            query = """
                UPDATE students 
                SET code = %s, name = %s, birth_date = %s, gender = %s, phone = %s, email = %s,
                    address = %s, interests = %s, education = %s, introduction = %s, campus = %s,
                    course_code = %s, notes = %s, photo_path = %s
                WHERE code = %s
            """
            self.db.execute_query(query, (
                code, name, birth_date, gender, phone, email,
                address, interests, education, introduction, campus, course_code, notes, self.current_photo_path,
                self.original_code
            ))
            
            QMessageBox.information(self, "성공", "학생 정보가 수정되었습니다.")
            self.clear_form()
            self.load_data()
            
        except Exception as e:
            error_msg = f"수정 실패: {str(e)}\n\n상세 오류:\n{traceback.format_exc()}"
            print(error_msg)
            QMessageBox.critical(self, "오류", error_msg)
            
    def delete_student(self):
        """학생 삭제"""
        code = self.code_input.text().strip()
        
        if not code:
            QMessageBox.warning(self, "경고", "삭제할 학생을 선택하세요.")
            return
        
        reply = QMessageBox.question(
            self, "확인", 
            f"학생 {code}를 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                query = "DELETE FROM students WHERE code = %s"
                self.db.execute_query(query, (code,))
                QMessageBox.information(self, "성공", "학생이 삭제되었습니다.")
                self.clear_form()
                self.load_data()
            except Exception as e:
                error_msg = f"삭제 실패: {str(e)}"
                print(error_msg)
                QMessageBox.critical(self, "오류", error_msg)
                
    def load_data(self):
        """데이터 로드"""
        try:
            if not self.db.connect():
                return
            
            query = """
                SELECT s.*, c.name as course_name
                FROM students s
                LEFT JOIN courses c ON s.course_code = c.code
                ORDER BY s.code
            """
            students = self.db.fetch_all(query)
            
            self.table.setRowCount(0)
            
            if students:
                self.table.setRowCount(len(students))
                for row_position, student in enumerate(students):
                    self.table.setItem(row_position, 0, QTableWidgetItem(student['code'] or ''))
                    self.table.setItem(row_position, 1, QTableWidgetItem(student['name'] or ''))
                    self.table.setItem(row_position, 2, QTableWidgetItem(student['birth_date'] or ''))
                    self.table.setItem(row_position, 3, QTableWidgetItem(student['gender'] or ''))
                    self.table.setItem(row_position, 4, QTableWidgetItem(student['phone'] or ''))
                    self.table.setItem(row_position, 5, QTableWidgetItem(student['email'] or ''))
                    self.table.setItem(row_position, 6, QTableWidgetItem(student['interests'] or ''))
                    self.table.setItem(row_position, 7, QTableWidgetItem(student['education'] or ''))
                    self.table.setItem(row_position, 8, QTableWidgetItem(student['campus'] or ''))
                    
                    course_display = student['course_name'] or '미배정'
                    self.table.setItem(row_position, 9, QTableWidgetItem(course_display))
                    
                    registered_at = student.get('registered_at')
                    if registered_at:
                        reg_str = registered_at.strftime('%Y-%m-%d') if isinstance(registered_at, datetime) else str(registered_at)
                        self.table.setItem(row_position, 10, QTableWidgetItem(reg_str))
                    else:
                        self.table.setItem(row_position, 10, QTableWidgetItem(''))
                    
                    self.table.setItem(row_position, 11, QTableWidgetItem(student['notes'] or ''))
                    
        except Exception as e:
            error_msg = f"데이터 로드 실패: {str(e)}\n\n상세 오류:\n{traceback.format_exc()}"
            print(error_msg)
            QMessageBox.critical(self, "오류", error_msg)
            
    def on_row_selected(self, row, column):
        """행 선택 시"""
        code = self.table.item(row, 0).text()
        self.original_code = code
        
        self.code_input.setText(code)
        self.name_input.setText(self.table.item(row, 1).text() if self.table.item(row, 1) else '')
        self.birth_input.setText(self.table.item(row, 2).text() if self.table.item(row, 2) else '')
        
        gender = self.table.item(row, 3).text() if self.table.item(row, 3) else ''
        gender_index = self.gender_combo.findText(gender)
        if gender_index >= 0:
            self.gender_combo.setCurrentIndex(gender_index)
        
        self.phone_input.setText(self.table.item(row, 4).text() if self.table.item(row, 4) else '')
        self.email_input.setText(self.table.item(row, 5).text() if self.table.item(row, 5) else '')
        self.interests_input.setText(self.table.item(row, 6).text() if self.table.item(row, 6) else '')
        self.education_input.setText(self.table.item(row, 7).text() if self.table.item(row, 7) else '')
        self.campus_input.setText(self.table.item(row, 8).text() if self.table.item(row, 8) else '')
        
        # DB에서 상세 정보 조회
        try:
            if not self.db.connect():
                return
            
            query = "SELECT * FROM students WHERE code = %s"
            student = self.db.fetch_one(query, (code,))
            
            if student:
                self.address_input.setText(student.get('address') or '')
                self.introduction_input.setText(student.get('introduction') or '')
                self.notes_input.setText(student.get('notes') or '')
                
                # 배정과정 설정
                course_code = student.get('course_code')
                if course_code:
                    for i in range(self.course_combo.count()):
                        if self.course_combo.itemData(i) == course_code:
                            self.course_combo.setCurrentIndex(i)
                            break
                else:
                    self.course_combo.setCurrentIndex(0)  # 미배정
                
                # 사진 로드
                self.load_student_photo(code)
                    
        except Exception as e:
            print(f"상세 정보 로드 오류: {str(e)}")
            
    def clear_form(self):
        """폼 초기화"""
        self.code_input.clear()
        self.name_input.clear()
        self.birth_input.clear()
        self.gender_combo.setCurrentIndex(0)
        self.phone_input.clear()
        self.email_input.clear()
        self.address_input.clear()
        self.interests_input.clear()
        self.education_input.clear()
        self.introduction_input.clear()
        self.campus_input.clear()
        self.course_combo.setCurrentIndex(0)
        self.notes_input.clear()
        self.original_code = None
        self.current_photo_path = None
        self.photo_label.clear()
        self.photo_label.setText("사진 없음")
    
    def upload_photo(self):
        """학생 사진 업로드"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "학생 사진 선택", "",
            "Image Files (*.png *.jpg *.jpeg *.bmp *.gif)"
        )
        
        if not file_path:
            return
        
        try:
            # 사진 저장 디렉토리 생성
            os.makedirs(PHOTO_ORIGINALS, exist_ok=True)
            os.makedirs(PHOTO_THUMBNAILS, exist_ok=True)
            
            # 파일명 생성 (학생 코드_timestamp.확장자)
            code = self.code_input.text().strip()
            if not code:
                QMessageBox.warning(self, "경고", "먼저 학생 코드를 입력하세요.")
                return
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_ext = os.path.splitext(file_path)[1]
            filename = f"{code}_{timestamp}{file_ext}"
            
            # 원본 저장
            original_path = os.path.join(PHOTO_ORIGINALS, filename)
            img = Image.open(file_path)
            img.save(original_path, quality=95)
            
            # 썸네일 생성 (150x180)
            thumbnail_path = os.path.join(PHOTO_THUMBNAILS, filename)
            img_copy = img.copy()
            img_copy.thumbnail((150, 180), Image.Resampling.LANCZOS)
            
            # 배경이 있는 썸네일 생성 (정확한 크기)
            thumbnail = Image.new('RGB', (150, 180), (255, 255, 255))
            offset = ((150 - img_copy.width) // 2, (180 - img_copy.height) // 2)
            thumbnail.paste(img_copy, offset)
            thumbnail.save(thumbnail_path, quality=85)
            
            # 현재 사진 경로 저장
            self.current_photo_path = original_path
            
            # 미리보기 표시
            pixmap = QPixmap(thumbnail_path)
            self.photo_label.setPixmap(pixmap)
            
            # DB에 사진 경로 저장 (학생이 이미 등록된 경우)
            if self.original_code:
                self.save_photo_to_db(self.original_code, original_path)
                QMessageBox.information(self, "성공", "사진이 등록되었습니다.")
            else:
                QMessageBox.information(self, "성공", 
                    "사진이 선택되었습니다.\n학생 추가/수정 시 자동으로 저장됩니다.")
            
        except Exception as e:
            QMessageBox.critical(self, "오류", f"사진 업로드 실패:\n{str(e)}")
    
    def remove_photo(self):
        """학생 사진 삭제"""
        if not self.current_photo_path and not self.original_code:
            QMessageBox.warning(self, "경고", "삭제할 사진이 없습니다.")
            return
        
        reply = QMessageBox.question(
            self, "확인",
            "사진을 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                # DB에서 사진 경로 삭제
                if self.original_code:
                    self.save_photo_to_db(self.original_code, None)
                
                # 파일 삭제 (실제로는 보관)
                self.current_photo_path = None
                self.photo_label.clear()
                self.photo_label.setText("사진 없음")
                
                QMessageBox.information(self, "성공", "사진이 삭제되었습니다.")
                
            except Exception as e:
                QMessageBox.critical(self, "오류", f"사진 삭제 실패:\n{str(e)}")
    
    def save_photo_to_db(self, student_code, photo_path):
        """DB에 사진 경로 저장"""
        try:
            if not self.db.connect():
                return
            
            query = "UPDATE students SET photo_path = %s WHERE code = %s"
            self.db.execute_query(query, (photo_path, student_code))
            
        except Exception as e:
            print(f"사진 경로 저장 오류: {str(e)}")
    
    def load_student_photo(self, student_code):
        """학생 사진 로드"""
        try:
            if not self.db.connect():
                return
            
            query = "SELECT photo_path FROM students WHERE code = %s"
            result = self.db.fetch_one(query, (student_code,))
            
            if result and result.get('photo_path'):
                photo_path = result['photo_path']
                self.current_photo_path = photo_path
                
                # 썸네일 경로 생성
                filename = os.path.basename(photo_path)
                thumbnail_path = os.path.join(PHOTO_THUMBNAILS, filename)
                
                # 썸네일 표시
                if os.path.exists(thumbnail_path):
                    pixmap = QPixmap(thumbnail_path)
                    self.photo_label.setPixmap(pixmap)
                elif os.path.exists(photo_path):
                    # 원본만 있으면 원본 표시
                    pixmap = QPixmap(photo_path)
                    scaled_pixmap = pixmap.scaled(150, 180, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.photo_label.setPixmap(scaled_pixmap)
                else:
                    self.photo_label.setText("사진 없음")
            else:
                self.current_photo_path = None
                self.photo_label.clear()
                self.photo_label.setText("사진 없음")
                
        except Exception as e:
            print(f"사진 로드 오류: {str(e)}")
