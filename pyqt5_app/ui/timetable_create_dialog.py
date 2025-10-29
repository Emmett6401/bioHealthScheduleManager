# -*- coding: utf-8 -*-
"""
시간표 작성 다이얼로그
"""

from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QComboBox, QMessageBox, QGroupBox,
                             QTableWidget, QTableWidgetItem, QHeaderView,
                             QProgressBar, QFrame, QScrollArea, QDialog,
                             QDialogButtonBox, QFileDialog, QInputDialog)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor, QBrush, QFont
from datetime import datetime, timedelta, time
import random
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database.db_manager import DatabaseManager


class TimetableCreateDialog(QWidget):
    """시간표 작성 위젯"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.db = DatabaseManager()
        self.selected_course = None
        self.subjects = []
        self.holidays = set()
        self.subject_colors = {}
        self.current_timetable = []  # 현재 표시 중인 시간표
        self.timetable_id = None  # 저장된 시간표 ID
        self.init_ui()
        self.load_courses()
        
    def init_ui(self):
        """UI 초기화"""
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # 과정 선택
        course_group = QGroupBox("📚 과정 선택")
        course_group.setStyleSheet("QGroupBox { font-size: 11pt; font-weight: bold; padding-top: 10px; }")
        course_layout = QHBoxLayout()
        
        self.course_combo = QComboBox()
        self.course_combo.setMinimumHeight(32)
        self.course_combo.setStyleSheet("font-size: 11pt;")
        self.course_combo.currentIndexChanged.connect(self.on_course_selected)
        course_layout.addWidget(self.course_combo)
        
        self.course_info_label = QLabel("과정을 선택하세요")
        self.course_info_label.setStyleSheet("font-size: 11pt; color: #666;")
        course_layout.addWidget(self.course_info_label)
        course_layout.addStretch()
        
        course_group.setLayout(course_layout)
        layout.addWidget(course_group)
        
        # 과목 목록
        subject_group = QGroupBox("📋 과목 목록")
        subject_group.setStyleSheet("QGroupBox { font-size: 11pt; font-weight: bold; padding-top: 10px; }")
        subject_layout = QVBoxLayout()
        
        self.subject_table = QTableWidget()
        self.subject_table.setColumnCount(7)
        self.subject_table.setHorizontalHeaderLabels([
            "과목명", "시수", "일수", "주강사", "보조강사", "예비강사", "색상"
        ])
        self.subject_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        # 6개 행이 보이도록 높이 설정 (헤더 30px + 행 6개 * 30px = 210px)
        self.subject_table.setFixedHeight(210)
        self.subject_table.setStyleSheet("font-size: 11pt;")
        subject_layout.addWidget(self.subject_table)
        
        subject_group.setLayout(subject_layout)
        layout.addWidget(subject_group)
        
        # 자동 배정 버튼
        btn_layout = QHBoxLayout()
        
        self.auto_btn = QPushButton("🎯 자동 배정")
        self.auto_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px 20px; font-size: 11pt;")
        self.auto_btn.setMinimumHeight(40)
        self.auto_btn.clicked.connect(self.auto_assign)
        self.auto_btn.setEnabled(False)
        btn_layout.addWidget(self.auto_btn)
        
        self.save_btn = QPushButton("💾 저장")
        self.save_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 10px 20px; font-size: 11pt;")
        self.save_btn.setMinimumHeight(40)
        self.save_btn.clicked.connect(self.save_timetable)
        self.save_btn.setEnabled(False)
        btn_layout.addWidget(self.save_btn)
        
        self.export_btn = QPushButton("📥 Excel 다운로드")
        self.export_btn.setStyleSheet("padding: 10px 20px; font-size: 11pt;")
        self.export_btn.setMinimumHeight(40)
        self.export_btn.clicked.connect(self.export_excel)
        self.export_btn.setEnabled(False)
        btn_layout.addWidget(self.export_btn)
        
        self.import_btn = QPushButton("📤 Excel 업로드")
        self.import_btn.setStyleSheet("padding: 10px 20px; font-size: 11pt;")
        self.import_btn.setMinimumHeight(40)
        self.import_btn.clicked.connect(self.import_excel)
        btn_layout.addWidget(self.import_btn)
        
        self.delete_btn = QPushButton("🗑️ 삭제")
        self.delete_btn.setStyleSheet("background-color: #F44336; color: white; padding: 10px 20px; font-size: 11pt;")
        self.delete_btn.setMinimumHeight(40)
        self.delete_btn.clicked.connect(self.delete_timetable)
        self.delete_btn.setEnabled(False)
        btn_layout.addWidget(self.delete_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 진행률 표시
        self.progress = QProgressBar()
        self.progress.setStyleSheet("font-size: 11pt;")
        self.progress.setVisible(False)
        layout.addWidget(self.progress)
        
        # 시간표 테이블
        timetable_group = QGroupBox("📅 생성된 시간표")
        timetable_group.setStyleSheet("QGroupBox { font-size: 11pt; font-weight: bold; padding-top: 10px; }")
        timetable_layout = QVBoxLayout()
        
        # 스크롤 영역
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMinimumHeight(400)
        
        self.timetable_table = QTableWidget()
        self.timetable_table.setColumnCount(8)
        self.timetable_table.setHorizontalHeaderLabels([
            "주차", "날짜", "오전(09:00-13:00)", "오후(14:00-18:00)", "주강사", "보조강사", "예비강사", "진행도"
        ])
        self.timetable_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.timetable_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.timetable_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.timetable_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.timetable_table.setStyleSheet("""
            QTableWidget {
                font-size: 11pt;
                gridline-color: #CCCCCC;
            }
        """)
        self.timetable_table.setShowGrid(True)
        self.timetable_table.cellClicked.connect(self.on_cell_clicked)
        
        scroll.setWidget(self.timetable_table)
        timetable_layout.addWidget(scroll)
        
        timetable_group.setLayout(timetable_layout)
        layout.addWidget(timetable_group)
        
        self.setLayout(layout)
    
    def load_courses(self):
        """과정 목록 로드"""
        try:
            if self.db.connect():
                query = "SELECT code, name, start_date FROM courses ORDER BY start_date DESC"
                courses = self.db.fetch_all(query)
                
                self.course_combo.clear()
                self.course_combo.addItem("-- 과정 선택 --", None)
                
                for course in courses:
                    display_text = f"{course['code']} - {course['name']}"
                    if course.get('start_date'):
                        display_text += f" ({course['start_date'].strftime('%Y-%m-%d')})"
                    self.course_combo.addItem(display_text, course['code'])
                    
        except Exception as e:
            QMessageBox.critical(self, "오류", f"과정 목록 로드 실패: {str(e)}")
    
    def on_course_selected(self, index):
        """과정 선택 시"""
        if index <= 0:
            self.selected_course = None
            self.auto_btn.setEnabled(False)
            self.course_info_label.setText("과정을 선택하세요")
            return
        
        course_code = self.course_combo.itemData(index)
        self.selected_course = course_code
        self.load_course_info()
        self.load_subjects()
        self.load_holidays()
        self.auto_btn.setEnabled(True)
    
    def load_course_info(self):
        """과정 정보 로드"""
        try:
            query = """
                SELECT start_date, lecture_end_date, lecture_hours 
                FROM courses 
                WHERE code = %s
            """
            result = self.db.fetch_one(query, (self.selected_course,))
            
            if result:
                info = f"시작: {result['start_date'].strftime('%Y-%m-%d')}"
                if result.get('lecture_end_date'):
                    info += f" ~ {result['lecture_end_date'].strftime('%Y-%m-%d')}"
                info += f" | 총 {result['lecture_hours']}시간"
                self.course_info_label.setText(info)
                
        except Exception as e:
            print(f"과정 정보 로드 오류: {str(e)}")
    
    def load_subjects(self):
        """과목 목록 로드"""
        try:
            query = """
                SELECT s.code, s.name, s.hours, 
                       s.day_of_week, s.is_biweekly, s.week_offset,
                       i1.name as main_instructor_name,
                       i2.name as assistant_instructor_name,
                       i3.name as reserve_instructor_name
                FROM subjects s
                LEFT JOIN instructors i1 ON s.main_instructor = i1.code
                LEFT JOIN instructors i2 ON s.assistant_instructor = i2.code
                LEFT JOIN instructors i3 ON s.reserve_instructor = i3.code
                ORDER BY s.hours ASC
            """
            self.subjects = self.db.fetch_all(query)
            
            self.subject_table.setRowCount(len(self.subjects))
            
            # 과목별 색상 생성
            self.subject_colors = {}
            colors = self.generate_colors(len(self.subjects))
            print(f"\n🎨 과목 색상 생성 시작 (총 {len(self.subjects)}개 과목)")
            
            for i, subject in enumerate(self.subjects):
                # 과목명
                self.subject_table.setItem(i, 0, QTableWidgetItem(subject['name']))
                
                # 시수
                hours = subject['hours']
                self.subject_table.setItem(i, 1, QTableWidgetItem(f"{hours}시간"))
                
                # 일수 (1일 8시간 기준, 소수점 1자리)
                days = hours / 8.0
                self.subject_table.setItem(i, 2, QTableWidgetItem(f"{days:.1f}일"))
                
                # 주강사
                main_instructor = subject.get('main_instructor_name') or '-'
                self.subject_table.setItem(i, 3, QTableWidgetItem(main_instructor))
                
                # 보조강사
                assistant_instructor = subject.get('assistant_instructor_name') or '-'
                self.subject_table.setItem(i, 4, QTableWidgetItem(assistant_instructor))
                
                # 예비강사
                reserve_instructor = subject.get('reserve_instructor_name') or '-'
                self.subject_table.setItem(i, 5, QTableWidgetItem(reserve_instructor))
                
                # 색상 (파스텔 톤)
                color = colors[i]
                self.subject_colors[subject['code']] = color
                color_item = QTableWidgetItem()
                color_item.setBackground(QBrush(color))
                self.subject_table.setItem(i, 6, color_item)
                
                # 디버깅: 색상 할당 확인
                print(f"  ✓ {subject['code']}: {subject['name'][:15]:15} → RGB({color.red():3}, {color.green():3}, {color.blue():3})")
                
        except Exception as e:
            QMessageBox.critical(self, "오류", f"과목 로드 실패: {str(e)}")
    
    def load_holidays(self):
        """공휴일 로드"""
        try:
            query = "SELECT holiday_date FROM holidays"
            rows = self.db.fetch_all(query)
            self.holidays = set([row['holiday_date'] for row in rows])
        except Exception as e:
            print(f"공휴일 로드 오류: {str(e)}")
    
    def generate_colors(self, count):
        """과목별 고유 파스텔 색상 생성"""
        colors = []
        hue_step = 360 / count
        
        for i in range(count):
            hue = int(i * hue_step)
            # 파스텔 색상 (채도 40%, 밝기 95%)
            color = QColor.fromHsv(hue, int(255 * 0.4), int(255 * 0.95))
            colors.append(color)
        
        return colors
    
    def auto_assign(self):
        """자동 시간표 배정"""
        if not self.selected_course or not self.subjects:
            QMessageBox.warning(self, "경고", "과정과 과목을 먼저 선택하세요.")
            return
        
        try:
            self.progress.setVisible(True)
            self.progress.setValue(0)
            
            # 과정 정보 가져오기
            query = "SELECT start_date, lecture_end_date FROM courses WHERE code = %s"
            course = self.db.fetch_one(query, (self.selected_course,))
            
            if not course or not course.get('start_date'):
                QMessageBox.warning(self, "경고", "과정 시작일이 설정되지 않았습니다.")
                return
            
            start_date = course['start_date']
            end_date = course.get('lecture_end_date') or start_date + timedelta(days=100)
            
            # 시간표 생성
            self.current_timetable = self.create_timetable(start_date, end_date)
            
            # 테이블에 표시
            self.display_timetable(self.current_timetable)
            
            self.progress.setValue(100)
            self.progress.setVisible(False)
            
            self.save_btn.setEnabled(True)
            self.export_btn.setEnabled(True)
            
            QMessageBox.information(self, "완료", "시간표가 자동으로 생성되었습니다.")
            
        except Exception as e:
            self.progress.setVisible(False)
            QMessageBox.critical(self, "오류", f"시간표 생성 실패: {str(e)}")
    
    def _make_subject_entry(self, subject, hours):
        """과목 엔트리 생성 헬퍼"""
        return {
            'code': subject['code'],
            'name': subject['name'],
            'total_hours': subject['hours'],
            'main_instructor': subject.get('main_instructor_name', '-'),
            'assistant_instructor': subject.get('assistant_instructor_name', '-'),
            'reserve_instructor': subject.get('reserve_instructor_name', '-'),
            'hours': hours
        }
    
    def _make_empty_entry(self):
        """빈 시간 엔트리 생성 헬퍼"""
        return {
            'code': '',
            'name': '-',
            'total_hours': 0,
            'main_instructor': '-',
            'assistant_instructor': '-',
            'reserve_instructor': '-',
            'hours': 0
        }
    
    def _find_most_remaining(self, remaining):
        """가장 시수 많이 남은 과목 찾기"""
        max_h = 0
        best = None
        for s in self.subjects:
            h = remaining.get(s['code'], 0)
            if h > max_h:
                max_h = h
                best = s
        return best
    
    def create_timetable(self, start_date, end_date):
        """시간표 생성 - 완전히 단순화된 요일 기반 배정
        
        원칙:
        1. 각 과목은 미리 지정된 요일(day_of_week)에만 배정
        2. 격주 과목은 해당 주차에만 배정
        3. 하루 = 8시간 (AM 4h + PM 4h), 1일 1과목 원칙
        4. 모든 시수를 소진할 때까지 반복
        
        데이터베이스 필드:
        - day_of_week: 0=월, 1=화, 2=수, 3=목, 4=금
        - is_biweekly: True=격주, False=매주
        - week_offset: 0=1주차, 1=2주차 (격주인 경우만 사용)
        """
        timetable = []
        remaining = {s['code']: s['hours'] for s in self.subjects}
        
        print("\n" + "="*80)
        print("🎯 시간표 생성 시작")
        print(f"📅 기간: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}")
        print(f"📚 과목 수: {len(self.subjects)}")
        print(f"⏰ 총 시수: {sum(remaining.values())}시간")
        
        # 과목별 배정 요일 출력
        day_names = ["월", "화", "수", "목", "금"]
        print("\n📋 과목별 배정 정보:")
        for s in self.subjects:
            day = s.get('day_of_week')
            day_str = day_names[day] if day is not None and 0 <= day <= 4 else "미설정"
            biweekly = "격주" if s.get('is_biweekly') else "매주"
            week = f"/{s.get('week_offset', 0)+1}주차" if s.get('is_biweekly') else ""
            print(f"  • {s['name']:<25} : {day_str}요일, {biweekly}{week}, {s['hours']}시간")
        print("="*80 + "\n")
        
        current_date = start_date
        day_count = 0
        
        while current_date <= end_date and any(h > 0 for h in remaining.values()):
            # 주말 스킵
            if current_date.weekday() >= 5:
                current_date += timedelta(days=1)
                continue
            
            # 공휴일 스킵
            if current_date in self.holidays:
                print(f"🎊 {current_date.strftime('%Y-%m-%d')} - 공휴일 (스킵)")
                current_date += timedelta(days=1)
                continue
            
            # 현재 요일과 주차
            weekday = current_date.weekday()  # 0=월, 1=화, 2=수, 3=목, 4=금
            week_number = (current_date - start_date).days // 7  # 0부터 시작
            
            # 이 날짜에 배정할 과목 찾기
            subject = None
            candidates = []
            
            # 1단계: 요일이 일치하고 시수가 남은 과목 찾기
            for s in self.subjects:
                # 요일 체크
                if s.get('day_of_week') != weekday:
                    continue
                
                # 시수 체크
                if remaining.get(s['code'], 0) <= 0:
                    continue
                
                # 격주 체크
                if s.get('is_biweekly', False):
                    # 격주 과목: 주차 확인
                    if week_number % 2 != s.get('week_offset', 0):
                        continue
                
                candidates.append(s)
            
            # 2단계: 후보 중에서 선택 (시수가 많이 남은 것 우선)
            if candidates:
                subject = max(candidates, key=lambda x: remaining.get(x['code'], 0))
            
            # 3단계: 요일 매칭이 없으면 → 남은 시수가 있는 아무 과목이나 배정 (미완료 과목 방지)
            if not subject:
                # 요일 상관없이 시수가 남은 과목 찾기
                any_remaining = []
                for s in self.subjects:
                    if remaining.get(s['code'], 0) > 0:
                        any_remaining.append(s)
                
                if any_remaining:
                    # 시수가 가장 많이 남은 과목 선택
                    subject = max(any_remaining, key=lambda x: remaining.get(x['code'], 0))
                    print(f"⚠️  {current_date.strftime('%Y-%m-%d')} ({day_names[weekday]}) - "
                          f"요일 미지정 과목 우선 배정: {subject['name']}")
                else:
                    # 정말 모든 과목이 끝남
                    print(f"✅ {current_date.strftime('%Y-%m-%d')} ({day_names[weekday]}) - 모든 과목 완료!")
                    break
            
            # AM 시간 배정 (4시간)
            code = subject['code']
            am_hours = min(4, remaining[code])
            am_subj = self._make_subject_entry(subject, am_hours) if am_hours > 0 else None
            remaining[code] -= am_hours
            
            # 디버깅: 과목 코드 확인
            if am_subj:
                print(f"📅 {current_date.strftime('%m-%d')} AM: {am_subj['code']} ({am_subj['name']}) - {am_hours}h")
            
            # PM 시간 배정 (4시간)
            # 1) 오전 과목이 끝났으면 (남은 시수 0) → 다른 과목 배정
            # 2) 오전 과목이 계속되면 → 같은 과목 배정
            pm_subject = None
            pm_hours = 0
            
            if remaining[code] > 0:
                # AM 과목이 계속됨
                pm_subject = subject
                pm_hours = min(4, remaining[code])
            else:
                # AM 과목 완료 → 남은 시수가 가장 많은 다른 과목 찾기
                other_candidates = []
                for s in self.subjects:
                    if s['code'] == code:  # 같은 과목 제외
                        continue
                    if remaining.get(s['code'], 0) <= 0:  # 시수 없는 과목 제외
                        continue
                    other_candidates.append(s)
                
                if other_candidates:
                    # 남은 시수가 가장 많은 과목 선택
                    pm_subject = max(other_candidates, key=lambda x: remaining.get(x['code'], 0))
                    pm_hours = min(4, remaining[pm_subject['code']])
                    print(f"  ⚡ 오전 과목 완료 → 오후는 {pm_subject['name']} 배정")
            
            # PM 과목이 있으면 배정
            pm_subj = None
            if pm_subject and pm_hours > 0:
                pm_subj = self._make_subject_entry(pm_subject, pm_hours)
                remaining[pm_subject['code']] -= pm_hours
                
                # 디버깅: 과목 코드 확인
                print(f"📅 {current_date.strftime('%m-%d')} PM: {pm_subj['code']} ({pm_subj['name']}) - {pm_hours}h")
            
            day_count += 1
            
            # 빈 슬롯 채우기
            if not am_subj:
                am_subj = self._make_empty_entry()
            if not pm_subj:
                pm_subj = self._make_empty_entry()
            
            timetable.append({
                'date': current_date, 
                'am_subject': am_subj, 
                'pm_subject': pm_subj
            })
            
            # 진행 상황 출력
            biweekly_mark = "🔄" if subject.get('is_biweekly') else "📅"
            if pm_subj and pm_subj.get('code') and am_subj.get('code') == pm_subj['code']:
                # 같은 과목
                print(f"{biweekly_mark} {current_date.strftime('%Y-%m-%d')} ({day_names[weekday]}) - "
                      f"{subject['name']:<25} : AM {am_hours}h + PM {pm_hours}h = {am_hours + pm_hours}h "
                      f"(남은 시수: {remaining.get(code, 0)}h)")
            elif pm_subj and pm_subj.get('code'):
                # 다른 과목
                am_code = am_subj.get('code', '')
                pm_code = pm_subj.get('code', '')
                print(f"{biweekly_mark} {current_date.strftime('%Y-%m-%d')} ({day_names[weekday]}) - "
                      f"AM: {am_subj.get('name', '-'):<20} {am_hours}h (남은: {remaining.get(am_code, 0)}h) | "
                      f"PM: {pm_subj.get('name', '-'):<20} {pm_hours}h (남은: {remaining.get(pm_code, 0)}h)")
            else:
                # AM만 있음
                am_code = am_subj.get('code', '')
                print(f"{biweekly_mark} {current_date.strftime('%Y-%m-%d')} ({day_names[weekday]}) - "
                      f"AM: {am_subj.get('name', '-'):<20} {am_hours}h (남은: {remaining.get(am_code, 0)}h) | PM: -")
            
            current_date += timedelta(days=1)
        
        # 최종 결과
        print("\n" + "="*80)
        print("✅ 시간표 생성 완료")
        print(f"📊 총 {day_count}일 배정")
        print(f"⏰ 배정된 시수: {sum(s['hours'] for s in self.subjects) - sum(remaining.values())}시간")
        
        # 미완료 과목 확인
        incomplete = {k: v for k, v in remaining.items() if v > 0}
        if incomplete:
            print("\n⚠️  미완료 과목:")
            for code, hours in incomplete.items():
                subject_name = next((s['name'] for s in self.subjects if s['code'] == code), code)
                print(f"  • {subject_name}: {hours}시간 남음")
        else:
            print("\n🎉 모든 과목 시수 배정 완료!")
        
        print("="*80 + "\n")
        
        return timetable
    
    def display_timetable(self, timetable):
        """시간표 테이블에 표시"""
        print(f"\n📊 시간표 표시 시작 (총 {len(timetable)}일)")
        print(f"🎨 사용 가능한 과목 색상: {len(self.subject_colors)}개")
        if self.subject_colors:
            print(f"   색상 딕셔너리 키: {list(self.subject_colors.keys())}")
        
        self.timetable_table.setRowCount(len(timetable))
        
        # 주차별 파스텔 오렌지 색상 팔레트
        week_colors = [
            QColor(255, 229, 204),  # 연한 오렌지 1
            QColor(255, 218, 185),  # 연한 오렌지 2
            QColor(255, 239, 213),  # 연한 피치
            QColor(255, 228, 196),  # 비스크
            QColor(255, 235, 205),  # 블랜치드 아몬드
        ]
        
        # 과목별 누적 시수 계산
        subject_accumulated = {}
        
        # 시작일 기준으로 주차 계산
        if timetable:
            start_date = timetable[0]['date']
        
        previous_week = None
        week_start_row = 0
        
        for i, entry in enumerate(timetable):
            am_subject = entry.get('am_subject', {})
            pm_subject = entry.get('pm_subject', {})
            
            # 주차 계산 (시작일 기준)
            current_date = entry['date']
            days_diff = (current_date - start_date).days
            week_number = (days_diff // 7) + 1
            
            # 주차가 바뀔 때 이전 주차 셀 병합
            if previous_week is not None and week_number != previous_week:
                # 이전 주차의 셀들을 병합
                if i - week_start_row > 1:  # 2개 이상의 행이 있을 때만 병합
                    self.timetable_table.setSpan(week_start_row, 0, i - week_start_row, 1)
                week_start_row = i
            
            # 주차별 배경색 선택 (순환)
            base_color = week_colors[(week_number - 1) % len(week_colors)]
            
            # 주차가 바뀌는 첫 행은 약간 더 진한 색으로 강조
            is_week_start = (previous_week is None or week_number != previous_week)
            if is_week_start:
                # 첫 행은 약간 더 진한 색
                week_bg_color = QColor(
                    max(base_color.red() - 20, 0),
                    max(base_color.green() - 20, 0),
                    max(base_color.blue() - 20, 0)
                )
            else:
                week_bg_color = base_color
            
            # 주차 표시 (첫 번째 행에만 설정, 병합 후 자동으로 표시됨)
            if is_week_start:
                week_item = QTableWidgetItem(f"{week_number}주차")
                week_item.setTextAlignment(Qt.AlignCenter)
                week_item.setBackground(QBrush(base_color))
                week_item.setFont(QFont("맑은 고딕", 11, QFont.Bold))
                self.timetable_table.setItem(i, 0, week_item)
            
            previous_week = week_number
            
            # 날짜
            date_str = entry['date'].strftime("%Y-%m-%d (%a)")
            date_item = QTableWidgetItem(date_str)
            date_item.setData(Qt.UserRole, entry)  # 데이터 저장
            date_item.setBackground(QBrush(week_bg_color))
            self.timetable_table.setItem(i, 1, date_item)
            
            # 오전 과목 표시
            if am_subject and am_subject.get('code'):
                am_code = am_subject['code']
                am_name = am_subject['name']
                am_total = am_subject['total_hours']
                am_hours_today = am_subject.get('hours', 0)
                
                # 누적 시수 계산 (오전)
                if am_code not in subject_accumulated:
                    subject_accumulated[am_code] = 0
                subject_accumulated[am_code] += am_hours_today
                am_accumulated = subject_accumulated[am_code]
                
                # 과목명 축약
                if len(am_name) > 10:
                    am_short = am_name[:10] + "..."
                else:
                    am_short = am_name
                
                # 누적 시수 표시
                am_text = f"{am_short}..({am_accumulated}h/{am_total}h)"
                am_item = QTableWidgetItem(am_text)
                am_item.setToolTip(f"{am_name}\n오늘 AM: {am_hours_today}h\n누적: {am_accumulated}h / {am_total}h")
                am_color = self.subject_colors.get(am_code, QColor(200, 200, 200))
                
                # 디버깅: 색상 적용 확인
                if am_code in self.subject_colors:
                    print(f"AM 색상 적용: {am_code} ({am_name}) → RGB({am_color.red()}, {am_color.green()}, {am_color.blue()})")
                else:
                    print(f"⚠️  AM 색상 없음: {am_code} ({am_name}) → 기본 회색 사용")
                
                am_item.setBackground(QBrush(am_color))
                am_item.setTextAlignment(Qt.AlignCenter)
                self.timetable_table.setItem(i, 2, am_item)
            else:
                empty_am = QTableWidgetItem("-")
                empty_am.setTextAlignment(Qt.AlignCenter)
                self.timetable_table.setItem(i, 2, empty_am)
            
            # 오후 과목 표시
            if pm_subject and pm_subject.get('code'):
                pm_code = pm_subject['code']
                pm_name = pm_subject['name']
                pm_total = pm_subject['total_hours']
                pm_hours_today = pm_subject.get('hours', 0)
                
                # 누적 시수 계산 (오후)
                if pm_code not in subject_accumulated:
                    subject_accumulated[pm_code] = 0
                subject_accumulated[pm_code] += pm_hours_today
                pm_accumulated = subject_accumulated[pm_code]
                
                # 과목명 축약
                if len(pm_name) > 10:
                    pm_short = pm_name[:10] + "..."
                else:
                    pm_short = pm_name
                
                # 누적 시수 표시
                pm_text = f"{pm_short}..({pm_accumulated}h/{pm_total}h)"
                pm_item = QTableWidgetItem(pm_text)
                pm_item.setToolTip(f"{pm_name}\n오늘 PM: {pm_hours_today}h\n누적: {pm_accumulated}h / {pm_total}h")
                pm_color = self.subject_colors.get(pm_code, QColor(200, 200, 200))
                
                # 디버깅: 색상 적용 확인
                if pm_code in self.subject_colors:
                    print(f"PM 색상 적용: {pm_code} ({pm_name}) → RGB({pm_color.red()}, {pm_color.green()}, {pm_color.blue()})")
                else:
                    print(f"⚠️  PM 색상 없음: {pm_code} ({pm_name}) → 기본 회색 사용")
                
                pm_item.setBackground(QBrush(pm_color))
                pm_item.setTextAlignment(Qt.AlignCenter)
                self.timetable_table.setItem(i, 3, pm_item)
            else:
                empty_pm = QTableWidgetItem("-")
                empty_pm.setTextAlignment(Qt.AlignCenter)
                self.timetable_table.setItem(i, 3, empty_pm)
            
            # 주강사 (오전 과목 기준, 오전/오후 다르면 둘 다 표시)
            if am_subject and pm_subject:
                if am_subject['code'] == pm_subject['code']:
                    instructor_text = am_subject.get('main_instructor', '-')
                else:
                    instructor_text = f"{am_subject.get('main_instructor', '-')} / {pm_subject.get('main_instructor', '-')}"
            elif am_subject:
                instructor_text = am_subject.get('main_instructor', '-')
            elif pm_subject:
                instructor_text = pm_subject.get('main_instructor', '-')
            else:
                instructor_text = '-'
            
            main_instructor_item = QTableWidgetItem(instructor_text)
            main_instructor_item.setTextAlignment(Qt.AlignCenter)
            main_instructor_item.setBackground(QBrush(week_bg_color))
            self.timetable_table.setItem(i, 4, main_instructor_item)
            
            # 보조강사
            if am_subject and pm_subject:
                if am_subject['code'] == pm_subject['code']:
                    assist_text = am_subject.get('assistant_instructor', '-')
                else:
                    assist_text = f"{am_subject.get('assistant_instructor', '-')} / {pm_subject.get('assistant_instructor', '-')}"
            elif am_subject:
                assist_text = am_subject.get('assistant_instructor', '-')
            elif pm_subject:
                assist_text = pm_subject.get('assistant_instructor', '-')
            else:
                assist_text = '-'
            
            assistant_instructor_item = QTableWidgetItem(assist_text)
            assistant_instructor_item.setTextAlignment(Qt.AlignCenter)
            assistant_instructor_item.setBackground(QBrush(week_bg_color))
            self.timetable_table.setItem(i, 5, assistant_instructor_item)
            
            # 예비강사
            if am_subject and pm_subject:
                if am_subject['code'] == pm_subject['code']:
                    reserve_text = am_subject.get('reserve_instructor', '-')
                else:
                    reserve_text = f"{am_subject.get('reserve_instructor', '-')} / {pm_subject.get('reserve_instructor', '-')}"
            elif am_subject:
                reserve_text = am_subject.get('reserve_instructor', '-')
            elif pm_subject:
                reserve_text = pm_subject.get('reserve_instructor', '-')
            else:
                reserve_text = '-'
            
            reserve_instructor_item = QTableWidgetItem(reserve_text)
            reserve_instructor_item.setTextAlignment(Qt.AlignCenter)
            reserve_instructor_item.setBackground(QBrush(week_bg_color))
            self.timetable_table.setItem(i, 6, reserve_instructor_item)
            
            # 진행도 표시 (오전 과목 기준)
            if am_subject:
                am_code = am_subject['code']
                am_total = am_subject['total_hours']
                am_accumulated = subject_accumulated.get(am_code, 0)
                progress_percent = (am_accumulated / am_total * 100) if am_total > 0 else 0
            else:
                progress_percent = 0
            
            progress_item = QTableWidgetItem(f"{progress_percent:.1f}%")
            progress_item.setTextAlignment(Qt.AlignCenter)
            progress_item.setBackground(QBrush(week_bg_color))
            self.timetable_table.setItem(i, 7, progress_item)
        
        # 마지막 주차 병합 처리
        if len(timetable) - week_start_row > 1:
            self.timetable_table.setSpan(week_start_row, 0, len(timetable) - week_start_row, 1)
    
    def on_cell_clicked(self, row, column):
        """셀 클릭 시 수정 가능"""
        if not self.current_timetable or row >= len(self.current_timetable):
            return
        
        entry = self.current_timetable[row]
        am_subject = entry.get('am_subject', {})
        pm_subject = entry.get('pm_subject', {})
        
        if column == 2:  # 오전 과목 - 과목 변경
            if not am_subject:
                return
            
            # 과목 선택 다이얼로그
            current_code = am_subject.get('code')
            dialog = SubjectSelectionDialog(self.subjects, current_code, self)
            
            if dialog.exec_() == QDialog.Accepted:
                new_subject_code = dialog.get_selected_subject()
                if new_subject_code and new_subject_code != current_code:
                    # 과목 정보 업데이트
                    new_subject = next((s for s in self.subjects if s['code'] == new_subject_code), None)
                    if new_subject:
                        # 오전 과목 교체 (시수는 유지)
                        am_subject['code'] = new_subject_code
                        am_subject['name'] = new_subject['name']
                        am_subject['total_hours'] = new_subject['hours']
                        am_subject['main_instructor'] = new_subject.get('main_instructor_name') or '-'
                        am_subject['assistant_instructor'] = new_subject.get('assistant_instructor_name') or '-'
                        am_subject['reserve_instructor'] = new_subject.get('reserve_instructor_name') or '-'
                        
                        # 테이블 전체 다시 표시 (진행도 재계산 필요)
                        self.display_timetable(self.current_timetable)
        
        elif column == 3:  # 오후 과목 - 과목 변경
            if not pm_subject:
                return
            
            # 과목 선택 다이얼로그
            current_code = pm_subject.get('code')
            dialog = SubjectSelectionDialog(self.subjects, current_code, self)
            
            if dialog.exec_() == QDialog.Accepted:
                new_subject_code = dialog.get_selected_subject()
                if new_subject_code and new_subject_code != current_code:
                    # 과목 정보 업데이트
                    new_subject = next((s for s in self.subjects if s['code'] == new_subject_code), None)
                    if new_subject:
                        # 오후 과목 교체 (시수는 유지)
                        pm_subject['code'] = new_subject_code
                        pm_subject['name'] = new_subject['name']
                        pm_subject['total_hours'] = new_subject['hours']
                        pm_subject['main_instructor'] = new_subject.get('main_instructor_name') or '-'
                        pm_subject['assistant_instructor'] = new_subject.get('assistant_instructor_name') or '-'
                        pm_subject['reserve_instructor'] = new_subject.get('reserve_instructor_name') or '-'
                        
                        # 테이블 전체 다시 표시 (진행도 재계산 필요)
                        self.display_timetable(self.current_timetable)
        
        elif column == 4:  # 주강사 클릭 - 예비강사와 교체
            # 오전/오후 같은 과목인지 확인
            if am_subject and pm_subject and am_subject.get('code') == pm_subject.get('code'):
                # 같은 과목 - 한 번만 교체
                main_inst = am_subject.get('main_instructor', '-')
                reserve_inst = am_subject.get('reserve_instructor', '-')
                
                if reserve_inst and reserve_inst != '-':
                    reply = QMessageBox.question(
                        self, "강사 교체", 
                        f"주강사와 예비강사를 교체하시겠습니까?\n\n주강사: {main_inst}\n예비강사: {reserve_inst}",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    
                    if reply == QMessageBox.Yes:
                        # 오전/오후 모두 교체
                        am_subject['main_instructor'], am_subject['reserve_instructor'] = \
                            am_subject['reserve_instructor'], am_subject['main_instructor']
                        pm_subject['main_instructor'], pm_subject['reserve_instructor'] = \
                            pm_subject['reserve_instructor'], pm_subject['main_instructor']
                        
                        # 테이블 업데이트
                        self.timetable_table.item(row, 4).setText(am_subject['main_instructor'])
                        self.timetable_table.item(row, 6).setText(am_subject['reserve_instructor'])
            else:
                # 다른 과목 - 둘 다 교체
                QMessageBox.information(self, "알림", 
                    "오전과 오후 과목이 다릅니다.\n각 과목의 셀을 따로 클릭하여 교체해주세요.")
        
        elif column == 5:  # 보조강사 클릭 - 수정
            # 오전/오후 같은 과목인지 확인
            if am_subject and pm_subject and am_subject.get('code') == pm_subject.get('code'):
                current_assistant = am_subject.get('assistant_instructor', '-')
            elif am_subject:
                current_assistant = am_subject.get('assistant_instructor', '-')
            else:
                current_assistant = '-'
            
            new_assistant, ok = QInputDialog.getText(
                self, "보조강사 수정", 
                "보조강사 이름을 입력하세요:",
                text=current_assistant if current_assistant != '-' else ''
            )
            
            if ok:
                new_value = new_assistant if new_assistant else '-'
                
                # 오전/오후 모두 업데이트
                if am_subject:
                    am_subject['assistant_instructor'] = new_value
                if pm_subject:
                    pm_subject['assistant_instructor'] = new_value
                
                self.timetable_table.item(row, 5).setText(new_value)
        
        elif column == 6:  # 예비강사 클릭 - 주강사와 교체
            # 오전/오후 같은 과목인지 확인
            if am_subject and pm_subject and am_subject.get('code') == pm_subject.get('code'):
                # 같은 과목 - 한 번만 교체
                main_inst = am_subject.get('main_instructor', '-')
                reserve_inst = am_subject.get('reserve_instructor', '-')
                
                if main_inst and main_inst != '-':
                    reply = QMessageBox.question(
                        self, "강사 교체", 
                        f"예비강사와 주강사를 교체하시겠습니까?\n\n예비강사: {reserve_inst}\n주강사: {main_inst}",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    
                    if reply == QMessageBox.Yes:
                        # 오전/오후 모두 교체
                        am_subject['reserve_instructor'], am_subject['main_instructor'] = \
                            am_subject['main_instructor'], am_subject['reserve_instructor']
                        pm_subject['reserve_instructor'], pm_subject['main_instructor'] = \
                            pm_subject['main_instructor'], pm_subject['reserve_instructor']
                        
                        # 테이블 업데이트
                        self.timetable_table.item(row, 4).setText(am_subject['main_instructor'])
                        self.timetable_table.item(row, 6).setText(am_subject['reserve_instructor'])
            else:
                # 다른 과목 - 둘 다 교체
                QMessageBox.information(self, "알림", 
                    "오전과 오후 과목이 다릅니다.\n각 과목의 셀을 따로 클릭하여 교체해주세요.")
    
    def save_timetable(self):
        """시간표 저장"""
        if not self.current_timetable:
            QMessageBox.warning(self, "경고", "저장할 시간표가 없습니다.")
            return
        
        try:
            # 기존 시간표 삭제 (있다면)
            delete_query = "DELETE FROM timetables WHERE course_code = %s AND type = 'lecture'"
            self.db.execute_query(delete_query, (self.selected_course,))
            
            # 새 시간표 저장
            insert_query = """
                INSERT INTO timetables 
                (course_code, subject_code, class_date, start_time, end_time, instructor_code, type)
                VALUES (%s, %s, %s, %s, %s, %s, 'lecture')
            """
            
            for entry in self.current_timetable:
                am_subject = entry.get('am_subject', {})
                pm_subject = entry.get('pm_subject', {})
                
                # 오전 시간표 저장
                if am_subject:
                    instructor_code = None
                    main_instructor = am_subject.get('main_instructor', '-')
                    if main_instructor and main_instructor != '-':
                        query = "SELECT code FROM instructors WHERE name = %s LIMIT 1"
                        result = self.db.fetch_one(query, (main_instructor,))
                        if result:
                            instructor_code = result['code']
                    
                    self.db.execute_query(insert_query, (
                        self.selected_course,
                        am_subject['code'],
                        entry['date'],
                        time(9, 0),
                        time(13, 0),
                        instructor_code
                    ))
                
                # 오후 시간표 저장
                if pm_subject:
                    instructor_code = None
                    main_instructor = pm_subject.get('main_instructor', '-')
                    if main_instructor and main_instructor != '-':
                        query = "SELECT code FROM instructors WHERE name = %s LIMIT 1"
                        result = self.db.fetch_one(query, (main_instructor,))
                        if result:
                            instructor_code = result['code']
                    
                    self.db.execute_query(insert_query, (
                        self.selected_course,
                        pm_subject['code'],
                        entry['date'],
                        time(14, 0),
                        time(18, 0),
                        instructor_code
                    ))
            
            self.delete_btn.setEnabled(True)
            QMessageBox.information(self, "완료", "시간표가 저장되었습니다.")
            
        except Exception as e:
            QMessageBox.critical(self, "오류", f"시간표 저장 실패: {str(e)}")
    
    def export_excel(self):
        """Excel 내보내기"""
        if not self.current_timetable:
            QMessageBox.warning(self, "경고", "내보낼 시간표가 없습니다.")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # 파일 저장 경로 선택
            file_path, _ = QFileDialog.getSaveFileName(
                self, "시간표 저장", f"시간표_{self.selected_course}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "시간표"
            
            # 헤더 작성
            headers = ["주차", "날짜", "요일", "오전(09:00-13:00)", "오후(14:00-18:00)", "주강사", "보조강사", "예비강사", "진행도"]
            ws.append(headers)
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 과목별 누적 시수 계산
            subject_accumulated = {}
            
            # 시작일 기준
            if self.current_timetable:
                start_date = self.current_timetable[0]['date']
            
            # 데이터 작성
            for entry in self.current_timetable:
                am_subject = entry.get('am_subject', {})
                pm_subject = entry.get('pm_subject', {})
                
                # 누적 시수 계산 (오전)
                if am_subject:
                    am_code = am_subject['code']
                    if am_code not in subject_accumulated:
                        subject_accumulated[am_code] = 0
                    subject_accumulated[am_code] += am_subject['hours']
                
                # 누적 시수 계산 (오후)
                if pm_subject:
                    pm_code = pm_subject['code']
                    if pm_code not in subject_accumulated:
                        subject_accumulated[pm_code] = 0
                    subject_accumulated[pm_code] += pm_subject['hours']
                
                # 주차 계산
                current_date = entry['date']
                days_diff = (current_date - start_date).days
                week_number = (days_diff // 7) + 1
                
                date_str = entry['date'].strftime("%Y-%m-%d")
                weekday = ["월", "화", "수", "목", "금", "토", "일"][entry['date'].weekday()]
                
                # 오전 과목 표시
                if am_subject:
                    am_name = am_subject['name']
                    am_total = am_subject['total_hours']
                    am_accumulated = subject_accumulated.get(am_subject['code'], 0)
                    am_text = f"{am_name}({am_accumulated}h/{am_total}h)"
                else:
                    am_text = "-"
                
                # 오후 과목 표시
                if pm_subject:
                    pm_name = pm_subject['name']
                    pm_total = pm_subject['total_hours']
                    pm_accumulated = subject_accumulated.get(pm_subject['code'], 0)
                    pm_text = f"{pm_name}({pm_accumulated}h/{pm_total}h)"
                else:
                    pm_text = "-"
                
                # 주강사 표시
                if am_subject and pm_subject:
                    if am_subject['code'] == pm_subject['code']:
                        main_inst = am_subject.get('main_instructor', '-')
                    else:
                        main_inst = f"{am_subject.get('main_instructor', '-')} / {pm_subject.get('main_instructor', '-')}"
                elif am_subject:
                    main_inst = am_subject.get('main_instructor', '-')
                elif pm_subject:
                    main_inst = pm_subject.get('main_instructor', '-')
                else:
                    main_inst = '-'
                
                # 보조강사 표시
                if am_subject and pm_subject:
                    if am_subject['code'] == pm_subject['code']:
                        assist_inst = am_subject.get('assistant_instructor', '-')
                    else:
                        assist_inst = f"{am_subject.get('assistant_instructor', '-')} / {pm_subject.get('assistant_instructor', '-')}"
                elif am_subject:
                    assist_inst = am_subject.get('assistant_instructor', '-')
                elif pm_subject:
                    assist_inst = pm_subject.get('assistant_instructor', '-')
                else:
                    assist_inst = '-'
                
                # 예비강사 표시
                if am_subject and pm_subject:
                    if am_subject['code'] == pm_subject['code']:
                        reserve_inst = am_subject.get('reserve_instructor', '-')
                    else:
                        reserve_inst = f"{am_subject.get('reserve_instructor', '-')} / {pm_subject.get('reserve_instructor', '-')}"
                elif am_subject:
                    reserve_inst = am_subject.get('reserve_instructor', '-')
                elif pm_subject:
                    reserve_inst = pm_subject.get('reserve_instructor', '-')
                else:
                    reserve_inst = '-'
                
                # 진행도 (오전 과목 기준)
                if am_subject:
                    progress_percent = (subject_accumulated.get(am_subject['code'], 0) / am_subject['total_hours'] * 100) if am_subject['total_hours'] > 0 else 0
                else:
                    progress_percent = 0
                
                row = [
                    f"{week_number}주차",
                    date_str,
                    weekday,
                    am_text,
                    pm_text,
                    main_inst,
                    assist_inst,
                    reserve_inst,
                    f"{progress_percent:.1f}%"
                ]
                ws.append(row)
                
                # 과목 색상 적용
                row_idx = ws.max_row
                
                # 오전 색상
                if am_subject:
                    am_color = self.subject_colors.get(am_subject['code'])
                    if am_color:
                        hex_color = "{:02X}{:02X}{:02X}".format(am_color.red(), am_color.green(), am_color.blue())
                        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        ws.cell(row_idx, 4).fill = fill
                
                # 오후 색상
                if pm_subject:
                    pm_color = self.subject_colors.get(pm_subject['code'])
                    if pm_color:
                        hex_color = "{:02X}{:02X}{:02X}".format(pm_color.red(), pm_color.green(), pm_color.blue())
                        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        ws.cell(row_idx, 5).fill = fill
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 10  # 주차
            ws.column_dimensions['B'].width = 15  # 날짜
            ws.column_dimensions['C'].width = 8   # 요일
            ws.column_dimensions['D'].width = 35  # 오전
            ws.column_dimensions['E'].width = 35  # 오후
            ws.column_dimensions['F'].width = 12  # 주강사
            ws.column_dimensions['G'].width = 12  # 보조강사
            ws.column_dimensions['H'].width = 12  # 예비강사
            ws.column_dimensions['I'].width = 10  # 진행도
            
            # 테두리 스타일
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=9):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 파일 저장
            wb.save(file_path)
            QMessageBox.information(self, "완료", f"시간표가 저장되었습니다.\n{file_path}")
            
        except ImportError:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 필요합니다.\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"Excel 내보내기 실패: {str(e)}")
    
    def import_excel(self):
        """Excel 가져오기"""
        try:
            from openpyxl import load_workbook
            
            # 파일 선택
            file_path, _ = QFileDialog.getOpenFileName(
                self, "시간표 불러오기", "", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Excel 파일 읽기
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 데이터 파싱
            self.current_timetable = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # 날짜가 없으면 스킵
                    continue
                
                date_str = str(row[0])
                subject_name = row[2]  # 오전 과목
                instructor_name = row[4] if len(row) > 4 else '-'
                
                # 날짜 파싱
                try:
                    if isinstance(row[0], datetime):
                        class_date = row[0].date()
                    else:
                        class_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                except:
                    continue
                
                # 과목 찾기
                subject = next((s for s in self.subjects if s['name'] == subject_name), None)
                if subject:
                    self.current_timetable.append({
                        'date': class_date,
                        'subject_code': subject['code'],
                        'subject_name': subject['name'],
                        'instructor': instructor_name,
                        'hours': 8
                    })
            
            # 테이블에 표시
            self.display_timetable(self.current_timetable)
            self.save_btn.setEnabled(True)
            self.export_btn.setEnabled(True)
            
            QMessageBox.information(self, "완료", f"{len(self.current_timetable)}개의 시간표 항목을 불러왔습니다.")
            
        except ImportError:
            QMessageBox.critical(self, "오류", "openpyxl 라이브러리가 필요합니다.\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"Excel 가져오기 실패: {str(e)}")
    
    def delete_timetable(self):
        """시간표 삭제"""
        reply = QMessageBox.question(
            self, "확인", "현재 과정의 시간표를 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                delete_query = "DELETE FROM timetables WHERE course_code = %s AND type = 'lecture'"
                self.db.execute_query(delete_query, (self.selected_course,))
                
                # 테이블 초기화
                self.timetable_table.setRowCount(0)
                self.current_timetable = []
                self.save_btn.setEnabled(False)
                self.export_btn.setEnabled(False)
                self.delete_btn.setEnabled(False)
                
                QMessageBox.information(self, "완료", "시간표가 삭제되었습니다.")
                
            except Exception as e:
                QMessageBox.critical(self, "오류", f"시간표 삭제 실패: {str(e)}")


class SubjectSelectionDialog(QDialog):
    """과목 선택 다이얼로그"""
    
    def __init__(self, subjects, current_subject_code, parent=None):
        super().__init__(parent)
        self.subjects = subjects
        self.current_subject_code = current_subject_code
        self.selected_subject_code = None
        self.init_ui()
    
    def init_ui(self):
        """UI 초기화"""
        self.setWindowTitle("과목 변경")
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        
        # 안내 메시지
        label = QLabel("변경할 과목을 선택하세요:")
        label.setStyleSheet("font-size: 11pt; font-weight: bold;")
        layout.addWidget(label)
        
        # 과목 콤보박스
        self.subject_combo = QComboBox()
        self.subject_combo.setStyleSheet("font-size: 11pt;")
        self.subject_combo.setMinimumHeight(32)
        
        for subject in self.subjects:
            self.subject_combo.addItem(
                f"{subject['name']} ({subject['hours']}시간)",
                subject['code']
            )
            if subject['code'] == self.current_subject_code:
                self.subject_combo.setCurrentIndex(self.subject_combo.count() - 1)
        
        layout.addWidget(self.subject_combo)
        
        # 버튼
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
        self.setLayout(layout)
    
    def accept(self):
        """확인 버튼"""
        self.selected_subject_code = self.subject_combo.currentData()
        super().accept()
    
    def get_selected_subject(self):
        """선택된 과목 코드 반환"""
        return self.selected_subject_code
